# -*- coding:utf-8 -*-
from argparse import ArgumentParser
from openpyxl import Workbook
import pandas as pd
import numpy as np
import glob
import os


def check_gender(gender_provided, gender):
    if gender_provided == 'unknown':
        result = 'NA'
    elif gender_provided == gender:
        result = 'PASS'
    else:
        result = 'NO PASS'
    return result


def check_cov(sample, coverage):
    if sample.startswith('Embryo') and coverage >= 0.7:
        result = 'PASS'
    elif not sample.startswith('Embryo') and coverage >= 0.9:
        result = 'PASS'
    else:
        result = 'NO PASS'
    return result


def check_mean_depth(sample, depth):
    if sample.startswith('Embryo') and depth >= 15:
        result = 'PASS'
    elif not sample.startswith('Embryo') and depth >= 15:
        result = 'PASS'
    else:
        result = 'NO PASS'
    return result


def final_qc(wgs_qc_file, lfr_qc_file):
    # LFR QC处理展示
    lfr_qc = pd.read_excel(lfr_qc_file, sheet_name='QC')
    lfr_qc['raw_bases'] = lfr_qc.apply(lambda x: f"{round(x['number_reads'] / np.power(10, 7))}G", axis=1)
    lfr_qc['coverage(rm_dup)'] = lfr_qc['coverage>=10(rm_dup)']
    lfr_qc['Gender_result'] = lfr_qc.apply(lambda x: check_gender(x['gender_provided'], x['gender']), axis=1)
    lfr_qc['cov_result'] = lfr_qc.apply(lambda x: check_cov(x['sample'], x['coverage(rm_dup)']), axis=1)
    lfr_qc['mean_depth_result'] = lfr_qc.apply(lambda x: check_mean_depth(x['sample'], x['mean_depth(rm_dup)']), axis=1)
    depth_cols = list(filter(lambda x: 'mean_depth(rm_dup)_' in x, lfr_qc.columns))
    cover_cols = list(filter(lambda x: 'coverage>=10(rm_dup)_' in x, lfr_qc.columns))
    lfr_out_cols = ['sample', 'sampleID', 'raw_bases', 'mean_depth(rm_dup)', 'coverage(rm_dup)'] + depth_cols + cover_cols + \
                   ['depX', 'depY', 'gender', 'Gender_result', 'cov_result', 'mean_depth_result']
    lfr_qc = lfr_qc[lfr_out_cols].copy()
    lfr_qc.rename(columns={col: col.replace('mean_depth(rm_dup)', 'gene_depth(rm_dup)') for col in depth_cols}, inplace=True)
    lfr_qc.rename(columns={col: col.replace('coverage>=10(rm_dup)', 'gene_coverage(rm_dup)') for col in cover_cols}, inplace=True)
    # wgs QC处理展示
    try:
        wgs_qc = pd.read_excel(wgs_qc_file)
        wgs_qc['coverage(rm_dup)'] = wgs_qc['coverage>=4(rm_dup)']
        wgs_qc['Gender_result'] = wgs_qc.apply(lambda x: check_gender(x['gender_provided'], x['gender']), axis=1)
        wgs_qc['cov_result'] = wgs_qc.apply(lambda x: check_cov(x['sample'], x['coverage(rm_dup)']), axis=1)
        wgs_qc['mean_depth_result'] = wgs_qc.apply(lambda x: check_mean_depth(x['sample'], x['mean_depth(rm_dup)']), axis=1)
        depth_cols = list(filter(lambda x: 'mean_depth(rm_dup)_' in x, wgs_qc.columns))
        cover_cols = list(filter(lambda x: 'coverage>=4(rm_dup)_' in x, wgs_qc.columns))
        ado_cols = list(filter(lambda x: 'ADO_' in x, wgs_qc.columns))
        wgs_out_cols = ['sample', 'sampleID', 'raw_bases', 'mean_depth(rm_dup)', 'coverage(rm_dup)'] + depth_cols + cover_cols + ado_cols + \
                       ['depX', 'depY', 'gender', 'Gender_result', 'cov_result', 'mean_depth_result']
        wgs_qc = wgs_qc[wgs_out_cols].copy()
        wgs_qc.rename(columns={col: col.replace('mean_depth(rm_dup)', 'gene_depth(rm_dup)') for col in depth_cols}, inplace=True)
        wgs_qc.rename(columns={col: col.replace('coverage>=4(rm_dup)', 'gene_coverage(rm_dup)') for col in cover_cols}, inplace=True)
    except KeyError:
        wgs_qc = pd.DataFrame()
    qc = wgs_qc.append(lfr_qc, sort=False)
    qc.fillna('.', inplace=True)
    return qc


def final_annotation(work_dir):
    # 合并变异注释结果
    annotation = pd.DataFrame()
    for anno in glob.glob(f'{work_dir}/Annotation/*.anno.tsv'):
        df = pd.read_csv(anno, sep='\t')
        annotation = annotation.append(df, sort=False)
    annotation.drop_duplicates(subset=['CHROM', 'POS', 'REF', 'ALT', 'sample', 'Transcript', 'HGVS.c'], inplace=True)
    annotation.sort_values(by=['CHROM', 'POS'], inplace=True)
    return annotation


def final_hap(work_dir, info):
    embryos = info[info['sample'].str.contains('Embryo')]['sample'].unique().tolist()
    # 简化输出hap变异信息
    hap_dic = dict()
    for h in glob.glob(f'{work_dir}/Hap/*.hap.xlsx'):
        df = pd.read_excel(h)
        # 过滤snp和IF
        df = df.loc[(df['VarType'] == 'snp') & (df['IF'] == 'yes')]
        if 'Maternal' in os.path.basename(h):
            cols = ['CHROM', 'POS', 'dbSNP', 'REF', 'ALT'] + [f'{s}_TGT' for s in ['Mother', 'Father'] + embryos] + \
                ['PS', 'TGT_lfr', 'M1', 'M2', 'Allele', 'IF_filter', 'Target', 'Location', 'Variant']
            hap_dic['Maternal'] = df[cols]
        elif 'Paternal' in os.path.basename(h):
            cols = ['CHROM', 'POS', 'dbSNP', 'REF', 'ALT'] + [f'{s}_TGT' for s in ['Mother', 'Father'] + embryos] + \
                ['PS', 'TGT_lfr', 'F1', 'F2', 'Allele', 'IF_filter', 'Target', 'Location', 'Variant']
            hap_dic['Paternal'] = df[cols]
        else:
            raise Exception('unexpect condition')
    return hap_dic


def report_result(args):
    # 合并结果
    qc = final_qc(args.wgs_qc, args.lfr_qc)
    annotation = final_annotation(args.work_dir)
    info = pd.read_csv(args.info, sep='\t')
    hap_dic = final_hap(args.work_dir, info)
    # excel输出qc
    wb = Workbook()
    ws = wb.create_sheet('PGT-M_result')
    row_num, col_num = qc.shape
    for k in range(col_num):
        ws.cell(1, k + 1).value = qc.columns[k]
    for i in range(row_num):
        for j in range(col_num):
            ws.cell(i + 2, j + 1).value = qc.iloc[i, j]
    # 分型结果
    start_row = row_num + 3
    for score in glob.glob(f'{args.work_dir}/Hap/*.score.xlsx'):
        score_df = pd.read_excel(score)
        for k, c in enumerate(score_df.columns):
            ws.cell(start_row, k + 1).value = c
        for i in range(score_df.shape[0]):
            for j in range(score_df.shape[1]):
                ws.cell(start_row + i + 1, j + 1).value = score_df.iloc[i, j]
        start_row = start_row + score_df.shape[0] + 3
    # 输出变异
    ws = wb.create_sheet(f'mut')
    for k, c in enumerate(annotation.columns):
        ws.cell(1, k + 1).value = c
    for i in range(annotation.shape[0]):
        for j in range(annotation.shape[1]):
            ws.cell(i + 2, j + 1).value = annotation.iloc[i, j]
    # 输出hap
    for parent in hap_dic:
        ws = wb.create_sheet(f'hap.{parent}')
        for k, c in enumerate(hap_dic[parent].columns):
            ws.cell(1, k + 1).value = c
        for i in range(hap_dic[parent].shape[0]):
            for j in range(hap_dic[parent].shape[1]):
                ws.cell(i + 2, j + 1).value = hap_dic[parent].iloc[i, j]
    del wb['Sheet']
    wb.save(f'{args.out}')


def main():
    parser = ArgumentParser()
    parser.add_argument('-wgs_qc', help='wgs qc file', required=True)
    parser.add_argument('-lfr_qc', help='lfr qc file', required=True)
    parser.add_argument('-work_dir', help='work dir', required=True)
    parser.add_argument('-out', help='output file', required=True)
    parser.add_argument('-info', help='sample info', required=True)
    parsed_args = parser.parse_args()
    report_result(parsed_args)


if __name__ == '__main__':
    main()
