# -*- coding:utf-8 -*-
from argparse import ArgumentParser
from openpyxl import load_workbook
from scipy import stats
import numpy as np


def beta_pdf(size=1e6, alpha=4, beta=100):
    p = np.arange(1, size+1)/size
    pdf = stats.beta(alpha, beta).pdf(p)
    return p, pdf


def cal_bf(p, pdf, x, n):
    bf = np.mean(stats.binom.pmf(x, n, (1+p)/2)*pdf)/np.mean(stats.binom.pmf(x, n, (1-p)/2)*pdf + 2*stats.binom.pmf(x, n, 1/2))
    return np.round(bf, 3)


def cal_chi2_p(a, b, c, d):
    cc = stats.chi2_contingency(np.array([[a, b], [c, d]]))
    return cc[1]


def cal_theta(mut_count, total):
    theta = mut_count/total
    return theta


def genotyping(p, theta, total, bf):
    threshold_p = 0.05
    theta_up = 0.52
    theta_low = 0.48
    threshold_total = 1000
    threshold_bf = 1
    if p < threshold_p:
        if theta > theta_up:
            genotype = 'Homozygous genotype (conclusive)'
        elif theta < theta_low:
            genotype = 'Wildtype genotype (conclusive)'
        else:
            if total >= threshold_total:
                if bf >= threshold_bf:
                    genotype = 'Homozygous genotype (conclusive)'
                else:
                    genotype = f'Probably wildtype genotype ({chr(952)}<50%) or retest ({chr(952)}>=50%)'
            else:
                genotype = f'Possibly wildtype genotype ({chr(952)}<50%) or homozygous ({chr(952)}>50%) genotype'
    else:
        if theta > theta_up:
            if total >= threshold_total:
                if bf >= threshold_bf:
                    genotype = 'Homozygous genotype (conclusive)'
                else:
                    genotype = 'Retest'
            else:
                genotype = 'Probably homozygous genotype'
        elif theta < theta_low:
            genotype = 'Wildtype or heterozygous genotype (conclusive)'
        else:
            if total >= threshold_total:
                if bf >= threshold_bf:
                    genotype = 'Probably homozygous genotype'
                else:
                    genotype = 'Possibly heterozygous genotype'
            else:
                genotype = 'Inconclusive'
    return genotype


def judge(theta):
    theta_up = 0.52
    theta_low = 0.48
    if theta < theta_low or theta > theta_up:
        return 'yes'
    else:
        return 'no'


def parsing(parsed_args):
    _p, _pdf = beta_pdf()
    wb = load_workbook(filename=parsed_args.ddpcr)
    sheets = wb.sheetnames
    for s in parsed_args.sheet.split(','):
        ddpcr = wb[sheets[int(s)-1]]
        for i in range(3, ddpcr.max_row+1):
            wild = ddpcr[f'C{i}'].value
            mut = ddpcr[f'D{i}'].value
            total = wild + mut
            wild_control = ddpcr[f'F{i}'].value
            mut_control = ddpcr[f'G{i}'].value
            # total_control = wild_control + mut_control
            theta = cal_theta(mut, total)
            p = cal_chi2_p(wild, mut, wild_control, mut_control)
            bf = cal_bf(_p, _pdf, int(mut), int(total))
            genotype = genotyping(p, theta, total, bf)
            ddpcr[f'I{i}'].value = theta
            ddpcr[f'J{i}'].value = p
            ddpcr[f'K{i}'].value = judge(theta)
            ddpcr[f'L{i}'].value = bf
            ddpcr[f'M{i}'].value = genotype
    wb.save(parsed_args.out)


def main():
    parser = ArgumentParser()
    parser.add_argument('-ddpcr', help='ddpcr result file', required=True)
    parser.add_argument('-out', help='out excel', required=True)
    parser.add_argument('-sheet', help='excel sheet to analyze', default='1', type=str)
    parsed_args = parser.parse_args()
    parsing(parsed_args)


if __name__ == '__main__':
    main()
