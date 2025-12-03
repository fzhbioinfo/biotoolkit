"""通用解读表格汇总
"""
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd


class WriteXlsx:
    """根据excel模板和单元格样式生成解读表格
    """
    def __init__(self, template, style):
        self.work_book = load_workbook(template)
        self.style = style

    def get_columns_index(self, work_sheet) -> dict:
        """获取工作表里表头的列序号

        :param work_sheet: 工作表
        :type work_sheet: worksheet
        :return: {'COLUMN_NAME': num}
        :rtype: dict
        """
        column_index = {work_sheet.cell(1, i).value: i for i in range(1, work_sheet.max_column + 1)}
        return column_index

    def write_json_to_sheet(self, sheet_name, file_name):
        """将json格式内容写入工作表

        :param sheet_name: 工作表名
        :type sheet_name: str
        :param file_name: json格式文件名
        :type file_name: str
        """
        work_sheet = self.work_book[sheet_name]
        column_index = self.get_columns_index(work_sheet)
        df = pd.read_json(file_name, orient='records', lines=True)
        for row in range(df.shape[0]):
            for col in df.columns:
                if col not in column_index:
                    continue
                if sheet_name in self.style and col in self.style[sheet_name]:
                    self.write_cell(work_sheet, row + 2, column_index[col], df.loc[row, col],
                                    self.style[sheet_name][col])
                else:
                    self.write_cell(work_sheet, row + 2, column_index[col], df.loc[row, col])

    @staticmethod
    def write_cell(work_sheet, row, col, value, style=None):
        """填写单元格内容

        :param work_sheet: 工作表
        :type work_sheet: worksheet
        :param row: 行号
        :type row: int
        :param col: 列号
        :type col: int
        :param value: 单元格的内容
        :type value: int|str
        :param style: 单元格样式, defaults to None
        :type style: dict, optional
        """
        if style is None:
            work_sheet.cell(row, col).value = value
        elif style['type'] == 'hyperlink':
            work_sheet.cell(row, col).hyperlink = value
        elif style['type'] == 'dropdown_list':
            work_sheet.cell(row, col).value = value
            validation = DataValidation(type='list', formula1=f'"{",".join(style["list"])}"',
                                        allow_blank=True,error='entry not in the list',
                                        errorTitle='Invalid Entry')
            work_sheet.add_data_validation(validation)
            validation.add(work_sheet.cell(row, col))
        elif style['type'] == 'hyperlink_split':
            if value in ['-', None, '', '.', ' '] or pd.isna(value):
                work_sheet.cell(row, col).value = value
            else:
                value, link = value.split('|')
                work_sheet.cell(row, col).value = value
                work_sheet.cell(row, col).hyperlink = link

    def write_list_to_sheet(self, sheet_name, file_name):
        """将列表格式内容写入工作表第一列, 在已存在的行后面追加

        :param sheet_name: 工作表名
        :type sheet_name: str
        :param file_name: .list后缀的文件名
        :type file_name: str
        """
        work_sheet = self.work_book[sheet_name]
        row = work_sheet.max_row + 2
        df = pd.read_csv(file_name, sep='\t', header=None)
        for value in df[0].values:
            self.write_cell(work_sheet, row, 1, value)
            row += 1

    def write_table_to_sheet(self, sheet_name, file_name):
        """将表格格式内容写入工作表, 不写表头

        :param sheet_name: 工作表名
        :type sheet_name: str
        :param file_name: 表格格式.txt后缀的文件名
        :type file_name: str
        """
        work_sheet = self.work_book[sheet_name]
        df = pd.read_csv(file_name, sep='\t', header=None)
        for row in range(df.shape[0]):
            for col in range(df.shape[1]):
                self.write_cell(work_sheet, row + 1, col + 1, df.loc[row, col])

    def save_book(self, file_name):
        """保存工作簿

        :param file_name: excel文件名
        :type file_name: str
        """
        self.work_book.save(file_name)
        self.work_book.close()
