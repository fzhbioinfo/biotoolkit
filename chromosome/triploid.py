from sklearn.mixture import GaussianMixture
from matplotlib import pyplot as plt
from sklearn.metrics import r2_score
from argparse import ArgumentParser
from collections import defaultdict
from multiprocessing import Pool
from functools import partial
from itertools import groupby
from openpyxl import Workbook
from scipy import stats
from io import BytesIO
from PIL import Image
import pandas as pd
import numpy as np
import logging
import allel
import re
import os

logger = logging.getLogger(__name__)
formater = logging.Formatter(
    '[%(asctime)s-%(filename)s:%(lineno)s-%(levelname)s] %(message)s', '%Y-%m-%d %H:%M:%S'
)
handler = logging.StreamHandler()
handler.setFormatter(formater)
logger.addHandler(handler)
logger.setLevel(logging.DEBUG)

# 关闭互动模式，防止在没有显示器的机器运行报错
plt.ioff()

# 读取列
VCF_FIELDS = ['variants/POS', 'variants/is_snp', 'calldata/GT', 'calldata/DP', 'calldata/AD']

# 资源目录
RESOURCES = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'resources')

# 读取区域，排除掉所有的异染色质区
REGIONS = os.path.join(RESOURCES, 'regions.bed')

# 峰的起始位置
MEAN_INIT = np.array([[0.33], [0.66]])

# 异常倍性
PLOIDY = {1: 'H', 3: 'T'}

# 读取区域
with open(REGIONS) as f:
    FIT_REGIONS = {
        contig: ['{}:{}-{}'.format(*region.split('\t')) for region in regions]
        for contig, regions in groupby(map(str.strip, f.readlines()), lambda r: r.split('\t')[0])
    }


class Chromosome:
    """
    读取一条染色体的数据并进行拟合
    """
    def __init__(self, chromosome, data):
        self.chromosome = chromosome
        # 变异数量
        self._variant_num = data['variants/POS'].size
        # 变异位置
        self.pos = data['variants/POS']
        # 变异总深度
        self.dp = data['calldata/DP'].flatten()
        # 变异深度
        self.ad = data['calldata/AD'].reshape(self._variant_num, -1)
        # 基因型
        self.gt = allel.GenotypeArray(data['calldata/GT']).is_het().reshape(self._variant_num)
        # 变异类型
        self.is_snp = data['variants/is_snp']
        self._fit_ratio_data = None
        self._double_peek = GaussianMixture()
        self._depth_mean, self._depth_scale = None, None

    @classmethod
    def fit_vcf(
            cls, vcf, sample, chromosome, min_depth=10, min_num=1000, covariance_type='full',
            tol=0.001, reg_covar=1e-08, max_iter=1000, n_init=10, init_params='kmeans',
            weights_init=None, means_init=MEAN_INIT, precisions_init=None, random_state=None,
            warm_start=False, verbose=0, verbose_interval=10
    ):
        """
        从VCF读取数据并进行拟合
        :param vcf: vcf文件
        :param sample: 样本编号
        :param chromosome: 染色体编号
        :param min_depth: 最低总深度
        :param min_num: 最低变异数量，低于该数量的染色体认为该染色体不存在
        :param covariance_type: 拟合类型，一般无需改动
        :param tol: 拟合精度，一般无需改动
        :param reg_covar: 拟合精度，一般无需改动
        :param max_iter: 最大迭代次数， 一般无需改动
        :param n_init: 初始化迭代次数，一般无需改动
        :param init_params: 初始化参数，一般无需改动
        :param weights_init: 每个峰的权重，一般无需改动
        :param means_init: 初始峰值，一般无需改动
        :param precisions_init: 初始标准差，一般无需改动
        :param random_state: 随机种子，一般无需改动
        :param warm_start: 初始值有问题时进行警告，一般无需改动
        :param verbose: 日志级别，一般无需改动
        :param verbose_interval: 日志间隔，一般无需改动
        :return:
        """
        # 读取并且筛选区域
        variants = filter(
            lambda v: v is not None, map(
                lambda r: allel.read_vcf(vcf, fields=VCF_FIELDS, samples=[sample], region=r),
                FIT_REGIONS[chromosome]
            )
        )
        # 合并数据
        data = defaultdict(list)
        for variant in variants:
            for field, value in variant.items():
                data[field].append(value)
        data = {field: np.concatenate(value) for field, value in data.items()}

        # 低于最低变数的染色体判定为拷贝数为0
        if len(data) == 0:
            logger.info(
                'Fit %s: variant_num=%s, ratio_r2_score=%s, ratio_separate=%s, depth_r2_score=%s, '
                'depth_mean=%s', chromosome, 0, 1, 1, 1, 0
            )
            return 0
        if data['variants/POS'].size < min_num:
            logger.info(
                'Fit %s: variant_num=%s, ratio_r2_score=%s, ratio_separate=%s, depth_r2_score=%s, '
                'depth_mean=%s', chromosome, data['variants/POS'].size, 1, 1, 1, 0
            )
            return data['variants/POS'].size
        # 读取数据并进行拟合
        chromosome = cls(chromosome, data)
        chromosome.fit_ratio(
            min_depth=min_depth, covariance_type=covariance_type, tol=tol, reg_covar=reg_covar,
            max_iter=max_iter, n_init=n_init, init_params=init_params, weights_init=weights_init,
            means_init=means_init, precisions_init=precisions_init, random_state=random_state,
            warm_start=warm_start, verbose=verbose, verbose_interval=verbose_interval
        )
        chromosome.fit_depth()
        logger.info(
            'Fit %s: variant_num=%s, ratio_r2_score=%s, ratio_separate=%s, depth_r2_score=%s, '
            'depth_mean=%s', chromosome.chromosome, len(chromosome), chromosome.ratio_r2_score,
            chromosome.ratio_separate, chromosome.depth_r2_score, chromosome.depth_mean
        )
        return chromosome

    def __len__(self):
        return self._variant_num

    def ratio_plot(self, ax=None, min_depth=10, **kwargs):
        """
        绘制ration图
        :param ax: 将图绘制到给定的坐标系
        :param min_depth: 最小深度
        :param kwargs: 其他绘图参数
        :return: 已绘图的坐标系
        """
        if ax is None:  # 未传入坐标系的时候创建坐标系
            fig, ax = plt.subplots(**kwargs)
        # 标记有效区域
        for region in FIT_REGIONS[self.chromosome]:
            start, end = map(int, re.findall(r':(\d+)-(\d+)', region).pop())
            ax.axvspan(start, end, alpha=0.8, facecolor='C8')
        # 计算相应数据
        dp_cutoff = self.dp > min_depth
        is_snp = self.is_snp
        pos = self.pos[dp_cutoff & is_snp]
        dp = self.dp[dp_cutoff & is_snp]
        ad = self.ad[dp_cutoff & is_snp]
        ratio = ad / dp.reshape(-1, 1)
        # 绘图
        ax.plot(pos, ratio, '.', alpha=0.1, c='C0', markeredgewidth=0)
        ax.set_ylim(0, 1)
        ax.set_xlabel('Position')
        ax.set_ylabel('Ratio')
        ax.set_title('SNP ratio in {}'.format(self.chromosome))
        return ax

    def fit_ratio(
            self, min_depth=10, covariance_type='full', tol=0.001, reg_covar=1e-08, max_iter=1000,
            n_init=10, init_params='kmeans', weights_init=None, means_init=MEAN_INIT,
            precisions_init=None, random_state=None, warm_start=False, verbose=0,
            verbose_interval=10
    ):
        """
        对 SNP ratio 进行拟合
        :param min_depth: 最低总深度
        :param covariance_type: 拟合类型，一般无需改动
        :param tol: 拟合精度，一般无需改动
        :param reg_covar: 拟合精度，一般无需改动
        :param max_iter: 最大迭代次数， 一般无需改动
        :param n_init: 初始化迭代次数，一般无需改动
        :param init_params: 初始化参数，一般无需改动
        :param weights_init: 每个峰的权重，一般无需改动
        :param means_init: 初始峰值，一般无需改动
        :param precisions_init: 初始标准差，一般无需改动
        :param random_state: 随机种子，一般无需改动
        :param warm_start: 初始值有问题时进行警告，一般无需改动
        :param verbose: 日志级别，一般无需改动
        :param verbose_interval: 日志间隔，一般无需改动
        """
        # filters
        dp_cutoff = self.dp > min_depth
        is_het = self.gt
        is_snp = self.is_snp

        # prepare fit data
        dp = self.dp[dp_cutoff & is_het & is_snp]
        ad = self.ad[dp_cutoff & is_het & is_snp]
        ratio = ad / dp.reshape(-1, 1)
        ratio = ratio[ratio > 0].reshape(-1, 1)

        if ratio.size == 0:
            return None

        # fit and return model
        self._double_peek = GaussianMixture(
            2, covariance_type=covariance_type, tol=tol, reg_covar=reg_covar, max_iter=max_iter,
            n_init=n_init, init_params=init_params, weights_init=weights_init,
            means_init=means_init, precisions_init=precisions_init, random_state=random_state,
            warm_start=warm_start, verbose=verbose, verbose_interval=verbose_interval
        )
        self._double_peek.fit(ratio)
        self._fit_ratio_data = ratio
        return self._double_peek

    def ratio_fit_plot(self, ax=None, **kwargs):
        """
        绘制 ratio 的拟合图
        :param ax: 将图绘制到给定的坐标系
        :param kwargs: 其他绘图参数
        :return: 已经绘图的坐标系
        """
        if ax is None:  # 未传入坐标系时创建新的坐标系
            fig, ax = plt.subplots(**kwargs)
        # 计算相关数据
        x = np.linspace(0, 1, 100).reshape(-1, 1)
        xs = np.column_stack(2 * [x])
        ys = stats.norm.pdf(
            xs, self._double_peek.means_.flatten(),
            np.sqrt(1 / self._double_peek.precisions_.flatten())
        ) * self._double_peek.weights_
        # 绘制柱状图
        ax.hist(self._fit_ratio_data.flatten(), bins=100, density=True, label='real')
        ax.fill(x, ys.sum(axis=1), c='green', alpha=0.2)
        # 绘制拟合曲线
        ax.plot(x, ys.sum(axis=1), c='green', label='fit')
        ax.plot(x, ys, '--', c='orange', label='cluster')
        ax.set_xlabel('Ratio')
        ax.set_ylabel('Probability Density')
        legends = {label: handel for handel, label in zip(*ax.get_legend_handles_labels())}
        legend_labels = ['real', 'fit', 'cluster']
        ax.legend([legends[label] for label in legend_labels], legend_labels)
        ax.set_title(
            'SNP ratio PDF in {}(R-square={})'.format(self.chromosome, self.ratio_r2_score)
        )
        return ax

    @property
    def ratio_separate(self):
        """
        返回双峰的距离
        """
        if self._fit_ratio_data is None:
            raise AttributeError("call .fit_ratio to fit ratio model first!")
        return np.abs(np.subtract.reduce(self._double_peek.means_.flatten()))

    @property
    def ratio_r2_score(self):
        """
        返回 R2 score
        """
        if self._fit_ratio_data is None:
            raise AttributeError("call .fit_ratio to fit ratio model first!")
        pdf, bins = np.histogram(self._fit_ratio_data.flatten(), bins=100, density=True)
        mid = (bins[1:] + bins[:-1]) / 2
        return r2_score(pdf, np.exp(self._double_peek.score_samples(mid.reshape(-1, 1))))

    def depth_plot(self, ax=None, max_depth=100, depth=None, **kwargs):
        """
        绘制深度图
        :param ax: 将图绘制到给定的坐标系
        :param max_depth: 最大深度
        :param depth: 平均单倍体深度
        :param kwargs: 其他绘图参数
        :return: 绘图后的坐标系
        """
        if ax is None:  # 未传入坐标系时创建新的坐标系
            fig, ax = plt.subplots(**kwargs)
        # 绘制有效区域
        for region in FIT_REGIONS[self.chromosome]:
            start, end = map(int, re.findall(r':(\d+)-(\d+)', region).pop())
            ax.axvspan(start, end, alpha=0.8, facecolor='C8')
        # 绘制深度
        ax.plot(self.pos, self.dp, '.', alpha=0.1, c='C0')
        if depth is not None:
            ax.hlines(
                np.arange(1, 4) * depth, *ax.get_xlim(), colors='C2', linestyles='dashed', zorder=10
            )
        ax.set_ylim(0, max_depth)
        ax.set_xlabel('Position')
        ax.set_ylabel('Depth')
        ax.set_title('SNP depth in {}'.format(self.chromosome))
        return ax

    def depth_ratio_plot(self, depth, max_depth=100, ax=None, window_size=1000000, **kwargs):
        """
        绘制深度图
        :param depth: 平均深度
        :param max_depth: 最大深度
        :param ax: 将图绘制到给定的坐标系
        :param window_size: 窗口大小
        :param kwargs: 其他绘图参数
        :return: 绘图后的坐标系
        """
        # 计算相关数据
        if depth is None:
            data = pd.DataFrame(
                [self.pos, self.dp, self.pos // window_size], index=['pos', 'dp', 'label']
            ).T.groupby('label').mean()
        else:
            data = pd.DataFrame(
                [self.pos, self.dp / depth, self.pos // window_size], index=['pos', 'dp', 'label']
            ).T.groupby('label').mean()
        if ax is None:  # 未传入坐标系时创建新的坐标系
            fig, ax = plt.subplots(**kwargs)
        # 绘制有效区域
        for region in FIT_REGIONS[self.chromosome]:
            start, end = map(int, re.findall(r':(\d+)-(\d+)', region).pop())
            ax.axvspan(start, end, alpha=0.8, facecolor='C8')
        # 绘制
        ax.plot(data['pos'], data['dp'], '.', c='C0')
        # 绘制深度比例尺
        if depth is not None:
            ax.hlines(np.arange(1, 4), *ax.get_xlim(), colors='C2', linestyles='dashed', zorder=10)
            ax.set_ylim(0, 4)
        else:
            ax.set_ylim(0, max_depth)
        ax.set_xlabel('Position')
        ax.set_ylabel('Depth')
        ax.set_title('SNP depth ratio in {}'.format(self.chromosome))
        return ax

    def fit_depth(self):
        """
        拟合深度
        """
        is_snp = self.is_snp
        dp = self.dp[is_snp]
        self._depth_mean, self._depth_scale = stats.norm.fit(dp)
        return self._depth_mean, self._depth_scale

    @property
    def depth_mean(self):
        if self._depth_mean is None:
            self.fit_depth()
        return self._depth_mean

    @property
    def depth_scale(self):
        if self._depth_scale is None:
            self.fit_depth()
        return self._depth_scale

    @property
    def depth_r2_score(self):
        """
        计算深度拟合的 R2 Score
        """
        is_snp = self.is_snp
        dp = self.dp[is_snp]
        pdf, bins = np.histogram(dp, bins=100, density=True, range=(0, 100))
        mid = (bins[1:] + bins[:-1]) / 2
        return r2_score(pdf, stats.norm.pdf(mid, self.depth_mean, self.depth_scale))

    def get_image(self, min_depth, depth, max_depth):
        """
        绘制图片
        :param min_depth: 最低总深度
        :param depth: 平均深度
        :param max_depth: 最大深度
        :return: 图片
        """
        fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(8.27, 11.69))
        self.ratio_fit_plot(ax=ax1)
        self.ratio_plot(ax=ax2, min_depth=min_depth)
        self.depth_ratio_plot(ax=ax3, depth=depth, max_depth=max_depth)
        plt.tight_layout()
        buf = BytesIO()
        plt.savefig(buf, format='jpg', dpi=300)
        buf.seek(0)
        im = Image.open(buf)
        plt.close()
        return im


def write_excel(karyotype, sample, ploidy, filename):
    """
    输出表格数据
    :param karyotype: 核型
    :param sample: 样本编号
    :param ploidy: 拷贝数
    :param filename: 输出文件
    """
    wb = Workbook()
    ws = wb.active
    # 样本信息
    ws.append(['Sample', sample])
    ws.merge_cells('B1:E1')
    ws.append(['Karyotype', karyotype])
    ws.merge_cells('B2:E2')
    ws.append([])
    ws.merge_cells('A3:E3')
    # 详细信息
    ws.append(['Detail'])
    ws.merge_cells('A4:E4')
    ws.append(['Chromosome', 'Ploidy', 'Func', 'Score', 'R-square'])
    for i, (chrom, (p, score, r_square, func)) in enumerate(ploidy.items(), start=6):
        ws.append([chrom, p, func, score, r_square])
        if r_square < 0.7:
            for cell in sum(ws['A{}:E{}'.format(i, i)], tuple()):
                cell.style = 'Bad'
    logger.info('Karyotype of %s: %s', sample, karyotype)
    wb.save(filename)


def make_karyotype(ploidy):
    """
    根据拷贝数生成核型
    :param ploidy: 拷贝数
    :return: 核型
    """
    total = sum(p for p, *_ in ploidy.values())
    sex = ''.join(
        chrom.replace('chr', '') * p
        for chrom, (p, *_) in ploidy.items() if not re.match(r'chr\d+', chrom)
    )
    autosome = ''.join(
        '{}{}'.format(
            PLOIDY[p], chrom.replace('chr', '')
        ) for chrom, (p, *_) in ploidy.items() if p in PLOIDY and re.match(r'chr\d+', chrom)
    )
    if len(re.findall(r'T\d+', autosome)) == 22 or len(re.findall(r'H\d+', autosome)) == 22:
        autosome = ''
    karyotype = '{},{}{}'.format(total, sex, autosome)
    return karyotype


def make_plot(chromosomes, output, min_depth, max_depth, depth_mean=None):
    """
    对所有染色体进行绘图
    :param chromosomes: 染色体
    :param output: 输出文件
    :param min_depth: 最低深度
    :param max_depth: 最大深度
    :param depth_mean: 平均深度
    """
    ims = [
        chromosome.get_image(
            min_depth, depth_mean, max_depth
        ) for chromosome in chromosomes if not isinstance(chromosome, int)
    ]
    ims[0].save(output, "PDF", resolution=300.0, save_all=True, append_images=ims[1:])


def main(
        vcf, sample, output, processes=0, min_depth=10, min_num=1000, max_depth=100,
        ratio_r2_score=0.8, ratio_separate=0.3, covariance_type='full', tol=0.001, reg_covar=1e-08,
        max_iter=1000, n_init=10, init_params='kmeans', weights_init=None, means_init=MEAN_INIT,
        precisions_init=None, random_state=None, warm_start=False, verbose=0, verbose_interval=10
):
    # 判定最大可用核心数
    try:
        sys_max_process = os.environ['NUMEXPR_MAX_THREADS']
    except KeyError:
        sys_max_process = os.cpu_count()
    processes = min(sys_max_process, processes) if processes != 0 else sys_max_process
    # 启动进程分别拟合各个染色体
    wrapper = partial(
        Chromosome.fit_vcf, min_depth=min_depth, min_num=min_num, covariance_type=covariance_type,
        tol=tol, reg_covar=reg_covar, max_iter=max_iter, n_init=n_init, init_params=init_params,
        weights_init=weights_init, means_init=means_init, precisions_init=precisions_init,
        random_state=random_state, warm_start=warm_start, verbose=verbose,
        verbose_interval=verbose_interval
    )
    with Pool(processes) as pool:
        chromosomes = pool.starmap(
            wrapper, ((vcf, sample, chromosome) for chromosome in FIT_REGIONS.keys())
        )
    # 收集拟合结果
    ploidy, unknown_chromosomes, depth = {}, {}, []
    for chrom, chromosome in zip(FIT_REGIONS.keys(), chromosomes):
        if isinstance(chromosome, int):
            ploidy[chrom] = (0, chromosome, 1, 'variant_num')
        elif chromosome.ratio_r2_score > ratio_r2_score:
            if chromosome.ratio_separate > ratio_separate:
                ploidy[chrom] = (
                    3, chromosome.ratio_separate, chromosome.ratio_r2_score, 'ratio_separate'
                )
                depth.append(chromosome.dp[chromosome.is_snp] / 3)
            else:
                ploidy[chrom] = (
                    2, chromosome.ratio_separate, chromosome.ratio_r2_score, 'ratio_separate'
                )
                depth.append(chromosome.dp[chromosome.is_snp] / 2)
        else:
            unknown_chromosomes[chrom] = (chromosome.depth_mean, chromosome.depth_r2_score)

    # 无法判断
    if len(depth) == 0:
        make_plot(chromosomes, '{}.pdf'.format(output), min_depth, max_depth)
        logger.error('Sample {} no chromosome can be determine by ratio separate!'.format(sample))
        exit(0)

    # 整合结果并绘图
    depth = np.concatenate(depth)
    depth_mean, depth_scale = stats.norm.fit(depth)
    make_plot(chromosomes, '{}.pdf'.format(output), min_depth, max_depth, depth_mean)
    if len(unknown_chromosomes) > 0:
        pdf, bins = np.histogram(depth, bins=100, density=True, range=(0, 100))
        mid = (bins[1:] + bins[:-1]) / 2
        total_r2_score = r2_score(pdf, stats.norm.pdf(mid, depth_mean, depth_scale))
        for chrom, (depth, depth_r2_score) in unknown_chromosomes.items():
            score = depth / depth_mean
            ploidy[chrom] = (
                int(np.round(score)), score, total_r2_score * depth_r2_score, 'depth_ratio'
            )
    karyotype = make_karyotype(ploidy)
    # 输出结果
    write_excel(karyotype, sample, ploidy, '{}.xlsx'.format(output))


if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('sample', help='Sample in the VCF file')
    parser.add_argument('vcf', help='Path for the VCF file')
    parser.add_argument('output', help='output prefix')
    parser.add_argument('-d', type=int, help='minimum depth for a variant', default=10)
    parser.add_argument('-s', type=float, help='separate', default=0.3)
    parser.add_argument('-t', type=int, help='processes number', default=0)
    parser.add_argument('-m', type=int, help='max depth', default=100)
    parser.add_argument('-n', type=int, help='minimum number of variant in chromosome', default=500)
    parser.add_argument(
        '-r', type=float, help='minimum r-square score for fit goodness', default=0.8
    )
    parsed_args = parser.parse_args()
    logger.info('Start!')
    logger.info(parsed_args)
    main(
        parsed_args.vcf, parsed_args.sample, parsed_args.output, parsed_args.t,
        max_depth=parsed_args.m, ratio_separate=parsed_args.s, min_num=parsed_args.n,
        ratio_r2_score=parsed_args.r
    )
    logger.info('Finish!')

