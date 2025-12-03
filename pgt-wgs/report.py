# -*- coding:utf-8 -*-
from argparse import ArgumentParser
from openpyxl import Workbook
import pandas as pd
import glob
import os


def check_gender(gender_provided, gender):
    if gender_provided == 'unknown':
        result = 'NA'
    elif gender_provided == gender:
        result = 'PASS'
    else:
        result = 'NO PASS'
    return result


def check_cov(sample, coverage):
    if sample.startswith('Embryo') and coverage >= 0.7:
        result = 'PASS'
    elif not sample.startswith('Embryo') and coverage >= 0.9:
        result = 'PASS'
    else:
        result = 'NO PASS'
    return result


def check_rawbase(raw_bases):
    if int(raw_bases.replace('G', '')) >= 45:
        result = 'PASS'
    else:
        result = 'NO PASS'
    return result


def final_qc(args):
    qc_df = pd.DataFrame()
    for qc in glob.glob(f'{args.work_dir}/*/Summary/*.QC.tsv'):
        df = pd.read_csv(qc, sep='\t')
        qc_df = qc_df.append(df, sort=False)
    qc_df.drop_duplicates(subset=['sampleID'], inplace=True)
    qc_df['gender_result'] = qc_df.apply(lambda x: check_gender(x['gender_provided'], x['gender']), axis=1)
    qc_df['cov_result'] = qc_df.apply(lambda x: check_cov(x['sample'], x['coverage>=4(rm_dup)']), axis=1)
    qc_df['rawbase_result'] = qc_df.apply(lambda x: check_rawbase(x['raw_bases']), axis=1)
    return qc_df


def final_annotation(args):
    annotation = pd.DataFrame()
    for anno in glob.glob(f'{args.work_dir}/*/Summary/*.anno.tsv'):
        df = pd.read_csv(anno, sep='\t')
        annotation = annotation.append(df, sort=False)
    annotation.drop_duplicates(subset=['CHROM', 'POS', 'REF', 'ALT', 'sampleID', 'Transcript', 'HGVS.c'], inplace=True)
    annotation.sort_values(by=['CHROM', 'POS'], inplace=True)
    return annotation


def final_hap(args):
    info = pd.read_csv(args.info, sep='\t')
    info.drop_duplicates(subset=['sampleID'], inplace=True)
    info['ID'] = info['sample'] + '_' + info['sampleID']
    info_trio = pd.DataFrame()
    info_trio = info_trio.append(info[(info['sample'] == 'Paternal') & (info['relation'].str.contains('母亲'))], sort=False)
    info_trio = info_trio.append(info[(info['sample'] == 'Paternal') & (info['relation'].str.contains('父亲'))], sort=False)
    info_trio = info_trio.append(info[(info['sample'] == 'Paternal')], sort=False)
    info_trio = info_trio.append(info[(info['sample'] == 'Maternal') & (info['relation'].str.contains('母亲'))], sort=False)
    info_trio = info_trio.append(info[(info['sample'] == 'Maternal') & (info['relation'].str.contains('父亲'))], sort=False)
    info_trio = info_trio.append(info[(info['sample'] == 'Maternal')], sort=False)
    info_trio = info_trio.append(info[(info['sample'] == 'Proband')], sort=False)
    info_trio.drop_duplicates(subset=['sampleID'], inplace=True)
    info_trio = info_trio[info_trio['sample'].isin(['Paternal', 'Maternal', 'Proband'])].copy()
    score_dic = dict()
    for gene in args.gene.split(','):
        score_dic[gene] = list()
        for trio in info_trio['ID'].tolist():
            df = pd.read_excel(f'{args.work_dir}/{trio}/Summary/{trio}.{gene}.score.xlsx')
            relation = info_trio.loc[info_trio['ID'] == trio, 'relation'].values[0]
            for i in range(df.shape[0]):
                sample = df.loc[i, 'sample']
                if sample.startswith('Embryo'):
                    df.loc[i, 'sample'] = info.loc[info['sample'] == sample, 'ID'].values[0]
            df.rename(columns={'sample': trio, 'Gene': relation}, inplace=True)
            score_dic[gene].append(df)
    return score_dic


def analysis(args):
    # 获取胚胎列表
    info = pd.read_csv(args.info, sep='\t')
    embryos = info[info['sample'].str.contains('Embryo')]['sample'].unique().tolist()
    # 读取基因转录本对应关系
    gene_trans = pd.read_csv(args.trans, sep='\t', usecols=['gene', 'Gene Symbol', 'Transcript'])
    # 获取家系名称
    family = os.path.basename(args.work_dir)
    # 创建目录并拷贝数据
    if not os.path.exists(f'{args.work_dir}/{family}'):
        os.makedirs(f'{args.work_dir}/{family}')
    os.system(f'cp {args.work_dir}/*/Summary/*.hap.* {args.work_dir}/*/Summary/*.relationship.pdf {args.work_dir}/{family}')
    # 整理QC,变异注释,单体型结果
    qc = final_qc(args)
    qc.to_excel(f'{args.work_dir}/{family}/{family}.QC.xlsx', index=False)
    annotation = final_annotation(args)
    score_dic = final_hap(args)
    # 输出excel
    for gene in args.gene.split(','):
        # 输出QC
        qc_ = qc.copy()
        qc_.rename(columns={'ADO_all_gene': 'ADO', f'ADO_{gene}': 'ADO_gene', f'mean_depth(rm_dup)_{gene}': 'gene_depth'}, inplace=True)
        columns = ['sample', 'sampleID', 'raw_bases', 'mean_depth(rm_dup)', 'coverage>=4(rm_dup)', 'gene_depth', 'ADO', 'ADO_gene', 'depX',
                   'depY', 'gender', 'gender_result', 'cov_result', 'rawbase_result']
        qc_gene = qc_[columns].copy()
        wb = Workbook()
        ws = wb.create_sheet(gene)
        row_num, col_num = qc_gene.shape
        for k in range(col_num):
            ws.cell(1, k + 1).value = columns[k]
        for i in range(row_num):
            for j in range(col_num):
                ws.cell(i + 2, j + 1).value = qc_gene.iloc[i, j]
        # 输出单体型
        start_row = row_num + 3
        for df in score_dic[gene]:
            for k, c in enumerate(df.columns):
                ws.cell(start_row, k + 1).value = c
            for i in range(df.shape[0]):
                for j in range(df.shape[1]):
                    ws.cell(start_row + i + 1, j + 1).value = df.iloc[i, j]
            start_row = start_row + df.shape[0] + 3
        # 输出变异
        gene_trans_ = gene_trans[gene_trans['gene'] == gene]
        annotation_gene = annotation[annotation['Gene Symbol'].isin(gene_trans_['Gene Symbol'].tolist())].copy()
        ws = wb.create_sheet(f'mut')
        for k, c in enumerate(annotation_gene.columns):
            ws.cell(1, k + 1).value = c
        for i in range(annotation_gene.shape[0]):
            for j in range(annotation_gene.shape[1]):
                ws.cell(i + 2, j + 1).value = annotation_gene.iloc[i, j]
        # 输出hap
        for hap in glob.glob(f'{args.work_dir}/*/Summary/*.{gene}.hap.xlsx'):
            trio = os.path.basename(hap).split('.')[0]
            sample = trio.split('_')[0]
            hap_df = pd.read_excel(hap)
            hap_df = hap_df[(hap_df['VarType'] == 'snp') & (hap_df['IF'] == 'yes')]
            hap_cols = ['CHROM', 'POS', 'dbSNP', 'REF', 'ALT'] + [f'{s}_origin' for s in ['Father', 'Mother', sample] + embryos] + \
                ['F1', 'F2', 'M1', 'M2', 'Allele', 'IF_filter', 'Location']
            hap_df_ = hap_df[hap_cols]
            ws = wb.create_sheet(f'hap.{trio}')
            for k, c in enumerate(hap_cols):
                ws.cell(1, k + 1).value = c
            for i in range(hap_df_.shape[0]):
                for j in range(hap_df_.shape[1]):
                    ws.cell(i + 2, j + 1).value = hap_df_.iloc[i, j]
        del wb['Sheet']
        wb.save(f'{args.work_dir}/{family}/{args.prefix}.{family}.{gene}.xlsx')


def main():
    parser = ArgumentParser()
    parser.add_argument('-work_dir', help='work dir', required=True)
    parser.add_argument('-info', help='sample info', required=True)
    parser.add_argument('-gene', help='gene name', required=True)
    parser.add_argument('-trans', help='gene trans', required=True)
    parser.add_argument('-prefix', help='out file name prefix', required=True)
    parsed_args = parser.parse_args()
    analysis(parsed_args)


if __name__ == '__main__':
    main()
